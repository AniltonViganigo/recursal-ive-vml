from openpyxl import load_workbook
import os

# Carrega Relatorio Analitico
wb = load_workbook("..\\..\\Data\\Template\\Template.xlsx", False)
ws = wb.active

# Inicializa lista de Reclamantes
lista_empregados = []

# Lista dos arquivos txt
lista_arquivos = os.listdir("..\\..\\Data\\Extratos\\Caixa\\DGI\\")

for arquivo in lista_arquivos:
    inscricao_empresa = ""
    nome_empresa = ""
    cod_empresa = ""
    uf = ""

    # Inicia leitura do arquivo.txt
    with open("..\\..\\Data\\Extratos\\Caixa\\DGI\\" + arquivo, "r+") as f:
        lines = f.readlines()
        lines = lines[2:-2]

        for idx, line in enumerate(lines):
            lista_pivo = []

            # Captura: RECLAMADA
            if line.upper().__contains__("NOME DA EMPRESA:"):
                if nome_empresa == "":
                    nome_empresa = line.split("NOME DA EMPRESA:")[1].strip()

            # Captura: CODIGO EMPRESA
            if line.upper().__contains__("CODIGO DA EMPRESA"):
                if cod_empresa == "":
                    cod_empresa = lines[idx + 1].split()[0].strip()

            # Captura lista de Reclamantes
            if line.__contains__("VARA TRABALHISTA"):
                lista = lines[idx - 1].split("  ")
                for i in lista:
                    i.strip()
                    if len(i) > 1:
                        lista_pivo.append(i.strip())
                for i2 in line.split():
                    if i2.isdigit():
                        lista_pivo.append(i2.strip())
                lista_empregados.append(lista_pivo)

# Preeche Relatório Analitico
for n in range(len(lista_empregados)):
    index_excel = n + 2

    # VARA
    ws["A{}".format(index_excel)] = lista_empregados[n][8]

    # PROCESSO LOCALIZADO
    ws["D{}".format(index_excel)] = lista_empregados[n][9]

    # RECLAMANTE
    ws["F{}".format(index_excel)] = lista_empregados[n][0]

    # RECLAMADA
    ws["E{}".format(index_excel)] = nome_empresa

    # CODIGO EMPRESA
    ws["AF{}".format(index_excel)] = cod_empresa

    # CODIGO EMPREGADO
    ws["AG{}".format(index_excel)] = lista_empregados[n][1]

    # DATA ADMISSAO
    ws["AH{}".format(index_excel)] = lista_empregados[n][2]

    # PIS
    ws["AJ{}".format(index_excel)] = lista_empregados[n][3]

    # VALOR
    ws["G{}".format(index_excel)] = lista_empregados[n][6]


wb.save("..\\..\\Data\\RELATORIO.xlsx")
